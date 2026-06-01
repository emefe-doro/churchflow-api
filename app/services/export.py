import csv
import io
from datetime import datetime

from sqlalchemy import and_, select
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import joinedload

from app.models.contact import Contact
from app.models.follow_up import FollowUpTask, CommunicationLog
from app.models.user import User


class ExportService:
    def __init__(self, db: AsyncSession):
        self.db = db

    async def export_contacts(
        self, church_id: int, category: str | None = None
    ) -> list[dict]:
        conditions = [Contact.church_id == church_id, Contact.deleted_at.is_(None)]
        if category:
            conditions.append(Contact.category == category)

        stmt = (
            select(Contact)
            .options(
                joinedload(Contact.assigned_worker),
                joinedload(Contact.branch),
                joinedload(Contact.service_unit),
            )
            .where(and_(*conditions))
            .order_by(Contact.created_at.desc())
        )
        result = await self.db.execute(stmt)
        contacts = list(result.unique().scalars().all())

        return [
            {
                "First Name": c.first_name,
                "Last Name": c.last_name,
                "Phone": c.phone,
                "Email": c.email or "",
                "Gender": c.gender or "",
                "Age Group": c.age_group or "",
                "Address": c.address or "",
                "Category": c.category.replace("_", " ").title(),
                "Source": c.source or "",
                "Status": c.status.replace("_", " ").title(),
                "Worker": c.assigned_worker.name if c.assigned_worker else "",
                "Branch": c.branch.name if c.branch else "",
                "Foundation Class": c.foundation_class_status or "",
                "Service Unit": c.service_unit.name if c.service_unit else "",
                "Created": c.created_at.strftime("%Y-%m-%d") if c.created_at else "",
            }
            for c in contacts
        ]

    async def export_follow_ups(self, church_id: int) -> list[dict]:
        stmt = (
            select(FollowUpTask)
            .join(Contact, FollowUpTask.contact_id == Contact.id)
            .options(
                joinedload(FollowUpTask.assigned_user),
                joinedload(FollowUpTask.contact),
            )
            .where(Contact.church_id == church_id, Contact.deleted_at.is_(None))
            .order_by(FollowUpTask.due_date.desc())
        )
        result = await self.db.execute(stmt)
        tasks = list(result.unique().scalars().all())

        return [
            {
                "Contact": f"{t.contact.first_name} {t.contact.last_name}" if t.contact else "",
                "Phone": t.contact.phone if t.contact else "",
                "Task Type": t.task_type.replace("_", " ").title(),
                "Priority": t.priority.title(),
                "Status": t.status.replace("_", " ").title(),
                "Assigned To": t.assigned_user.name if t.assigned_user else "",
                "Due Date": t.due_date.strftime("%Y-%m-%d") if t.due_date else "",
                "Completed": t.completed_at.strftime("%Y-%m-%d") if t.completed_at else "",
                "Notes": t.notes or "",
                "Created": t.created_at.strftime("%Y-%m-%d") if t.created_at else "",
            }
            for t in tasks
        ]

    async def export_communication_logs(self, church_id: int) -> list[dict]:
        stmt = (
            select(CommunicationLog)
            .join(Contact, CommunicationLog.contact_id == Contact.id)
            .options(
                joinedload(CommunicationLog.sender),
                joinedload(CommunicationLog.contact),
            )
            .where(Contact.church_id == church_id, Contact.deleted_at.is_(None))
            .order_by(CommunicationLog.created_at.desc())
        )
        result = await self.db.execute(stmt)
        logs = list(result.unique().scalars().all())

        return [
            {
                "Contact": f"{l.contact.first_name} {l.contact.last_name}" if l.contact else "",
                "Phone": l.contact.phone if l.contact else "",
                "Channel": l.channel.replace("_", " ").title(),
                "Outcome": l.outcome or "",
                "Status": l.status.title(),
                "Sent By": l.sender.name if l.sender else "",
                "Message": l.message or "",
                "Date": (l.sent_at or l.created_at).strftime("%Y-%m-%d %H:%M"),
            }
            for l in logs
        ]

    @staticmethod
    def to_csv(rows: list[dict]) -> io.BytesIO:
        output = io.BytesIO()
        if rows:
            writer = csv.DictWriter(output, fieldnames=list(rows[0].keys()))
            writer.writeheader()
            writer.writerows(rows)
        output.seek(0)
        return output

    @staticmethod
    def to_excel(rows: list[dict], sheet_name: str) -> io.BytesIO:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter

        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name

        if rows:
            headers = list(rows[0].keys())
            header_fill = PatternFill(start_color="1E40AF", end_color="1E40AF", fill_type="solid")
            header_font = Font(color="FFFFFF", bold=True, size=11)

            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_idx, value=header)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center")

            for row_idx, row in enumerate(rows, 2):
                for col_idx, header in enumerate(headers, 1):
                    ws.cell(row=row_idx, column=col_idx, value=str(row.get(header, "")))

            for col_idx in range(1, len(headers) + 1):
                ws.column_dimensions[get_column_letter(col_idx)].width = 20

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output

    @staticmethod
    def to_pdf(rows: list[dict], title: str) -> io.BytesIO:
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib import colors
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import inch
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer

        output = io.BytesIO()
        page_size = landscape(A4) if len(rows[0]) > 7 else A4
        doc = SimpleDocTemplate(output, pagesize=page_size, topMargin=0.5 * inch, bottomMargin=0.3 * inch)
        elements = []

        styles = getSampleStyleSheet()
        title_style = styles["Heading1"]
        title_style.alignment = 1
        elements.append(Paragraph(title, title_style))
        elements.append(Spacer(1, 12))

        if rows:
            headers = list(rows[0].keys())
            table_data = [headers] + [[str(row.get(h, "")) for h in headers] for row in rows]

            col_widths = None
            if len(headers) > 7:
                available = page_size[0] - inch
                col_widths = [available / len(headers)] * len(headers)

            table = Table(table_data, colWidths=col_widths, repeatRows=1)
            table.setStyle(TableStyle([
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1E40AF")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTSIZE", (0, 0), (-1, 0), 9),
                ("BOTTOMPADDING", (0, 0), (-1, 0), 8),
                ("FONTSIZE", (0, 1), (-1, -1), 8),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F3F4F6")]),
            ]))
            elements.append(table)

        doc.build(elements)
        output.seek(0)
        return output
